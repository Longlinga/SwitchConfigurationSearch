import os
import re
import openpyxl


def DataSeach(file_str):
    # 匹配规则
    ruler = [r'user-interface vty 0',
             r'user-interface con 0'
             ]

    ruler1 = [r'sysname .*',
              r'undo ftp s.* all-interface|undo FTP s.* all-interface',
              r'undo ftp ipv6 .* all-interface|undo FTP ipv6 .* all-interface',
              r'undo telnet server-source all-interface',
              r'undo telnet ipv6 server-source all-interface',
              r'acl \d{4} inbound',
              r'snmp-agent group v3.* acl \d{4}',
              r'authentication-mode password'
              ]

    with open(file_str, 'r', encoding='UTF-8') as f:
        data = [f.name]
        # 读取文件内容并分割成行
        content = f.read().splitlines()
        # 逐行匹配规则
        for ruler in ruler:
            found_match = False
            ruler_pattern = re.compile(ruler)
            for i, line in enumerate(content):
                if ruler_pattern.search(line):
                    found_match = True
                    # 获取匹配值直至匹配值为'#'
                    search = ''
                    for j in range(i, len(content)):
                        if content[j] == '#':
                            break
                        search += content[j] + '\n'
                    data.append(search)
            if not found_match:
                data.append('N/A')
        f.close()

        # 打开文件
    with open(file_str, 'r', encoding='UTF-8') as file:
        content = file.read()
        for ruler1 in ruler1:
            found_match = False
            ruler_pattern = re.compile(ruler1)
            if re.search(ruler1, content):
                found_match = True
                # 获取匹配值
                match = re.search(ruler1, content).group()
                data.append(match)
            if not found_match:
                data.append('N/A')
        file.close()

    print(file_str + '搜索完毕')
    # 将data写入excel
    if os.path.exists('output.xlsx'):
        # 如果文件已存在，加载已有的Excel文件
        sheet = openpyxl.load_workbook('output.xlsx')
        ws = sheet.active
    else:
        # 如果文件不存在，创建新的Excel文件
        sheet = openpyxl.Workbook()
        ws = sheet.active
        ws.title = 'test'
    max_row = ws.max_row
    for i in range(len(data)):
        ws.cell(row=max_row + 1, column=i + 1, value=data[i])
    # 保存Excel文件
    sheet.save('output.xlsx')


if __name__ == '__main__':
    File_path = input('请输入文件路径：')
    FileName = os.listdir(File_path)
    for FileName in FileName:
        file_str = File_path + '/' + FileName
        DataSeach(file_str)
